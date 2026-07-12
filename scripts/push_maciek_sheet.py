#!/usr/bin/env python3
"""
Push data/maciek_plan_full.xlsx (tab "Plan") -> Google Sheet, mirroring
values + background colors + bold, cell-for-cell. The xlsx is the source of truth.

Setup (raz):
    python3 -m pip install --user gspread google-auth openpyxl
    # klucz SA trzymaj POZA repo, np.:
    #   ~/.config/gcloud/maciek-sa.json
    # i udostępnij arkusz e-mailowi SA jako Edytor.

Użycie:
    export MACIEK_SHEET_ID=<id z URL arkusza>         # .../spreadsheets/d/<ID>/edit
    export MACIEK_SA_KEY=~/.config/gcloud/maciek-sa.json   # (opcjonalne — to jest default)
    python3 scripts/push_maciek_sheet.py
    # lub:
    python3 scripts/push_maciek_sheet.py <SHEET_ID> [KEY_PATH]
"""
import os
import sys
import openpyxl
import gspread
from google.oauth2.service_account import Credentials

XLSX = os.path.join(os.path.dirname(__file__), "..", "data", "maciek_plan_full.xlsx")
SRC_TAB = "Plan"
DST_TAB = "Plan"
SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]
DEFAULT_KEY = os.path.expanduser("~/.config/gcloud/maciek-sa.json")


def argb_to_rgb(argb):
    """openpyxl '00RRGGBB' / 'FFRRGGBB' -> {red,green,blue} floats 0..1.
    Zwraca None dla theme/indexed color (nie-hex)."""
    s = str(argb) if argb is not None else ""
    if len(s) < 6:
        return None
    h = s[-6:]
    if any(ch not in "0123456789abcdefABCDEF" for ch in h):
        return None
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    return {"red": r / 255, "green": g / 255, "blue": b / 255}


def main():
    sheet_id = (sys.argv[1] if len(sys.argv) > 1 else os.environ.get("MACIEK_SHEET_ID"))
    key_path = (sys.argv[2] if len(sys.argv) > 2
                else os.environ.get("MACIEK_SA_KEY", DEFAULT_KEY))
    key_path = os.path.expanduser(key_path)
    if not sheet_id:
        sys.exit("Brak SHEET_ID. Podaj argumentem lub w MACIEK_SHEET_ID.")
    if not os.path.exists(key_path):
        sys.exit(f"Brak klucza SA: {key_path}")

    creds = Credentials.from_service_account_file(key_path, scopes=SCOPES)
    gc = gspread.authorize(creds)
    sh = gc.open_by_key(sheet_id)

    wb = openpyxl.load_workbook(XLSX)
    ws = wb[SRC_TAB]
    nrows, ncols = ws.max_row, ws.max_column

    # docelowy tab
    try:
        dst = sh.worksheet(DST_TAB)
    except gspread.WorksheetNotFound:
        dst = sh.add_worksheet(title=DST_TAB, rows=nrows + 10, cols=ncols + 2)
    dst.resize(rows=max(nrows, 1), cols=max(ncols, 1))

    # zbuduj grid CellData (wartość + tło + bold + kolor tekstu)
    rows_payload = []
    for r in range(1, nrows + 1):
        cells = []
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            fmt = {}
            bg = argb_to_rgb(cell.fill.fgColor.rgb) if cell.fill and cell.fill.patternType else None
            if bg:
                fmt["backgroundColor"] = bg
            tf = {}
            if cell.font and cell.font.bold:
                tf["bold"] = True
            fg = argb_to_rgb(cell.font.color.rgb) if (cell.font and cell.font.color) else None
            if fg:
                tf["foregroundColor"] = fg
            if cell.font and cell.font.size:
                tf["fontSize"] = int(cell.font.size)
            if tf:
                fmt["textFormat"] = tf
            data = {}
            v = cell.value
            if v is not None and v != "":
                data["userEnteredValue"] = {"stringValue": str(v)}
            if fmt:
                data["userEnteredFormat"] = fmt
            cells.append(data)
        rows_payload.append({"values": cells})

    requests = [
        # tekst do góry-lewej, zawijanie
        {"repeatCell": {
            "range": {"sheetId": dst.id, "startRowIndex": 0, "endRowIndex": nrows,
                      "startColumnIndex": 0, "endColumnIndex": ncols},
            "cell": {"userEnteredFormat": {"verticalAlignment": "MIDDLE",
                                            "wrapStrategy": "CLIP"}},
            "fields": "userEnteredFormat.verticalAlignment,userEnteredFormat.wrapStrategy"}},
        {"updateCells": {
            "rows": rows_payload,
            "fields": "userEnteredValue,userEnteredFormat.backgroundColor,userEnteredFormat.textFormat",
            "start": {"sheetId": dst.id, "rowIndex": 0, "columnIndex": 0}}},
    ]
    # szerokości kolumn wg xlsx (units openpyxl ~ *7 px)
    for i in range(ncols):
        col = openpyxl.utils.get_column_letter(i + 1)
        w = ws.column_dimensions[col].width
        if w:
            requests.append({"updateDimensionProperties": {
                "range": {"sheetId": dst.id, "dimension": "COLUMNS",
                          "startIndex": i, "endIndex": i + 1},
                "properties": {"pixelSize": int(w * 7)}, "fields": "pixelSize"}})

    sh.batch_update({"requests": requests})
    print(f"OK — wypchnięto {nrows}×{ncols} do taba '{DST_TAB}' "
          f"(https://docs.google.com/spreadsheets/d/{sheet_id})")


if __name__ == "__main__":
    main()
